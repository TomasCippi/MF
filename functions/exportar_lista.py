import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as ImagenExcel
from PIL import Image as ImagenPIL

from functions.db import obtener_productos
from functions.paths import obtener_carpeta_imgs
from functions.logger import obtener_logger
from functions.config import obtener_email_vendedor

logger = obtener_logger()

# ---------- Colores editables ----------
COLOR_PRINCIPAL = "4472C4"        # azul, sin el '#' (formato que pide openpyxl)
COLOR_TEXTO_BLANCO = "FFFFFF"
COLOR_TEXTO_AZUL_OSCURO = "1F3864"
COLOR_FONDO_BLANCO = "FFFFFF"

# ---------- Textos editables ----------
NOMBRE_EMPRESA = "MF Distribuidora"

# ---------- Tamaños editables ----------
TAMANO_IMAGEN_PX = 70   # tamaño del cuadrado de imagen dentro de la celda
ALTURA_FILA_CON_IMAGEN = 55   # alto de fila (en puntos) cuando el producto tiene imagen
ALTURA_FILA_SIN_IMAGEN = 20   # alto de fila normal, sin imagen

# Bordes finos negros, reutilizados en todo el archivo
BORDE_FINO = Side(style="thin", color="000000")
BORDE_COMPLETO = Border(left=BORDE_FINO, right=BORDE_FINO, top=BORDE_FINO, bottom=BORDE_FINO)

# Archivos temporales creados durante la exportación, para borrarlos al final
_archivos_temporales = []


def exportar_stock_excel(ruta_destino):
    """
    Genera el archivo Excel con el listado de stock en 'ruta_destino'.
    Devuelve True si se generó correctamente, False si hubo un error.
    """
    global _archivos_temporales
    _archivos_temporales = []

    try:
        productos = obtener_productos()
        carpeta_imgs = obtener_carpeta_imgs()

        # Determina si al menos un producto tiene imagen propia (no default.png)
        hay_imagen = any(
            p.get("imagen") and p.get("imagen") != "default.png"
            for p in productos
        )

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Stock"

        # Columnas usadas: B=codigo, C=producto, D=cantidad_caja,
        # (E=imagen si hay_imagen), y precio en la última columna usada.
        col_codigo = "B"
        col_producto = "C"
        col_caja = "D"
        if hay_imagen:
            col_imagen = "E"
            col_precio = "F"
            ultima_columna = "F"
        else:
            col_imagen = None
            col_precio = "E"
            ultima_columna = "E"

        # ---------- Fila 1: columna A vacía, como espaciador cuadrado ----------
        hoja.column_dimensions["A"].width = 4
        hoja.row_dimensions[1].height = 22

        # ---------- Fila 2: bloque de título "MF Distribuidora" ----------
        hoja.merge_cells(f"B2:{ultima_columna}2")
        celda_titulo = hoja["B2"]
        celda_titulo.value = NOMBRE_EMPRESA
        celda_titulo.font = Font(size=22, bold=True, color=COLOR_TEXTO_BLANCO)
        celda_titulo.fill = PatternFill("solid", fgColor=COLOR_PRINCIPAL)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        hoja.row_dimensions[2].height = 45

        for col in range(2, hoja[f"{ultima_columna}2"].column + 1):
            celda = hoja.cell(row=2, column=col)
            celda.border = BORDE_COMPLETO
            celda.fill = PatternFill("solid", fgColor=COLOR_PRINCIPAL)

        # ---------- Fila 3: bloque Email (B:C) y bloque Fecha ----------
        # El bloque Fecha se ubica pegado al final de la tabla: D:E si no
        # hay columna de imagen, E:F si sí la hay.
        if hay_imagen:
            col_fecha_label = "E"
            col_fecha_valor = "F"
        else:
            col_fecha_label = "D"
            col_fecha_valor = "E"

        hoja["B3"] = "Email:"
        hoja["B3"].font = Font(bold=True)
        hoja["C3"] = obtener_email_vendedor()

        hoja[f"{col_fecha_label}3"] = "Fecha:"
        hoja[f"{col_fecha_label}3"].font = Font(bold=True)
        hoja[f"{col_fecha_valor}3"] = datetime.now().strftime("%d/%m/%Y")

        for celda_ref in ["B3", "C3", f"{col_fecha_label}3", f"{col_fecha_valor}3"]:
            hoja[celda_ref].border = BORDE_COMPLETO
            hoja[celda_ref].alignment = Alignment(horizontal="left", vertical="center")

        # Columna C ancha, porque después van las descripciones de producto
        hoja.column_dimensions["B"].width = 16
        hoja.column_dimensions["C"].width = 40
        hoja.column_dimensions["D"].width = 16
        if hay_imagen:
            hoja.column_dimensions["E"].width = 14
            hoja.column_dimensions["F"].width = 16
        else:
            hoja.column_dimensions["E"].width = 16

        # ---------- Fila 4: encabezados de la tabla ----------
        fila_headers = 4
        encabezados = {
            col_codigo: "Código",
            col_producto: "Producto",
            col_caja: "Cantidad por caja",
        }
        if hay_imagen:
            encabezados[col_imagen] = "Imagen"
        encabezados[col_precio] = "Precio"

        for col_letra, texto in encabezados.items():
            celda = hoja[f"{col_letra}{fila_headers}"]
            celda.value = texto
            celda.font = Font(bold=True, color=COLOR_TEXTO_BLANCO)
            celda.fill = PatternFill("solid", fgColor=COLOR_PRINCIPAL)
            celda.border = BORDE_COMPLETO
            celda.alignment = Alignment(horizontal="center", vertical="center")

        # ---------- Filas de datos (ordenadas alfabéticamente por descripción) ----------
        productos_ordenados = sorted(productos, key=lambda p: p.get("descripcion", "").lower())

        fila_actual = fila_headers + 1

        for producto in productos_ordenados:
            nombre_imagen = producto.get("imagen")
            tiene_imagen_propia = bool(nombre_imagen and nombre_imagen != "default.png")

            ruta_imagen = None
            if tiene_imagen_propia:
                posible_ruta = carpeta_imgs / nombre_imagen
                if posible_ruta.exists():
                    ruta_imagen = posible_ruta

            # La altura de la fila depende de si ESTE producto puntual
            # tiene una imagen real disponible, no del archivo en general.
            hoja.row_dimensions[fila_actual].height = (
                ALTURA_FILA_CON_IMAGEN if ruta_imagen else ALTURA_FILA_SIN_IMAGEN
            )

            # Código: azul oscuro, fondo blanco
            celda_codigo = hoja[f"{col_codigo}{fila_actual}"]
            celda_codigo.value = producto.get("codigo", "")
            celda_codigo.font = Font(bold=True, color=COLOR_TEXTO_AZUL_OSCURO)
            celda_codigo.fill = PatternFill("solid", fgColor=COLOR_FONDO_BLANCO)
            celda_codigo.border = BORDE_COMPLETO
            celda_codigo.alignment = Alignment(horizontal="center", vertical="center")

            # Producto (descripción)
            celda_producto = hoja[f"{col_producto}{fila_actual}"]
            celda_producto.value = producto.get("descripcion", "")
            celda_producto.border = BORDE_COMPLETO
            celda_producto.alignment = Alignment(horizontal="left", vertical="center")

            # Cantidad por caja: fondo blanco, letra color principal
            celda_caja = hoja[f"{col_caja}{fila_actual}"]
            celda_caja.value = producto.get("cantidad_caja", 0)
            celda_caja.font = Font(bold=True, color=COLOR_PRINCIPAL)
            celda_caja.fill = PatternFill("solid", fgColor=COLOR_FONDO_BLANCO)
            celda_caja.border = BORDE_COMPLETO
            celda_caja.alignment = Alignment(horizontal="center", vertical="center")

            # Imagen (si corresponde la columna en este archivo)
            if hay_imagen:
                celda_imagen = hoja[f"{col_imagen}{fila_actual}"]
                celda_imagen.border = BORDE_COMPLETO

                if ruta_imagen:
                    _insertar_imagen(
                        hoja, str(ruta_imagen), col_imagen, fila_actual,
                        ALTURA_FILA_CON_IMAGEN
                    )

            # Precio: fondo color principal, letra blanca, más grande
            celda_precio = hoja[f"{col_precio}{fila_actual}"]
            celda_precio.value = producto.get("precio", 0)
            celda_precio.number_format = '"$"#,##0.00'
            celda_precio.font = Font(bold=True, size=13, color=COLOR_TEXTO_BLANCO)
            celda_precio.fill = PatternFill("solid", fgColor=COLOR_PRINCIPAL)
            celda_precio.border = BORDE_COMPLETO
            celda_precio.alignment = Alignment(horizontal="center", vertical="center")

            fila_actual += 1

        libro.save(ruta_destino)
        logger.warning(f"Exportación de stock generada en '{ruta_destino}' ({len(productos)} productos).")
        return True

    except Exception as e:
        logger.error(f"Error al exportar stock a Excel: {e}")
        return False

    finally:
        _limpiar_temporales()


def _insertar_imagen(hoja, ruta_imagen, col_letra, fila, alto_fila_pt):
    """
    Redimensiona la imagen del producto a un cuadrado que ocupa casi
    toda la celda, y la centra tanto horizontal como verticalmente,
    usando el ancho de columna y alto de fila reales para calcular
    el offset de centrado en píxeles.
    """
    imagen_pil = ImagenPIL.open(ruta_imagen).convert("RGB")
    imagen_pil.thumbnail((TAMANO_IMAGEN_PX, TAMANO_IMAGEN_PX), ImagenPIL.LANCZOS)

    ruta_temporal = f"{ruta_imagen}.thumb_{fila}.png"
    imagen_pil.save(ruta_temporal)
    _archivos_temporales.append(ruta_temporal)

    imagen_excel = ImagenExcel(ruta_temporal)
    imagen_excel.width = TAMANO_IMAGEN_PX
    imagen_excel.height = TAMANO_IMAGEN_PX

    # --- Cálculo de centrado ---
    # Ancho de columna en openpyxl está en "caracteres", 1 carácter ≈ 7px.
    # Alto de fila está en puntos, 1 punto ≈ 1.333px.
    ancho_columna_px = hoja.column_dimensions[col_letra].width * 7
    alto_fila_px = alto_fila_pt * 1.333

    offset_x = max(0, (ancho_columna_px - TAMANO_IMAGEN_PX) / 2)
    offset_y = max(0, (alto_fila_px - TAMANO_IMAGEN_PX) / 2)

    imagen_excel.anchor = _crear_ancla_centrada(hoja, col_letra, fila, offset_x, offset_y)
    hoja.add_image(imagen_excel)


def _crear_ancla_centrada(hoja, col_letra, fila, offset_x_px, offset_y_px):
    """
    Crea un ancla de tipo 'oneCell' con desplazamiento en EMU (unidad
    que usa Excel internamente), para que la imagen quede centrada
    dentro de la celda en vez de pegada a la esquina superior izquierda.
    """
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker

    columna_idx = hoja[f"{col_letra}1"].column - 1  # índice 0-based
    fila_idx = fila - 1  # índice 0-based

    marcador = AnchorMarker(
        col=columna_idx, colOff=pixels_to_EMU(offset_x_px),
        row=fila_idx, rowOff=pixels_to_EMU(offset_y_px)
    )

    tamano = XDRPositiveSize2D(
        pixels_to_EMU(TAMANO_IMAGEN_PX), pixels_to_EMU(TAMANO_IMAGEN_PX)
    )

    return OneCellAnchor(_from=marcador, ext=tamano)


def _limpiar_temporales():
    """Borra los archivos .thumb_*.png generados durante la exportación."""
    for ruta in _archivos_temporales:
        try:
            if os.path.exists(ruta):
                os.remove(ruta)
        except Exception:
            pass