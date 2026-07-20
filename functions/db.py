import sqlite3
from openpyxl import load_workbook

from functions.paths import obtener_ruta_db, hacer_backup_db
from functions.logger import obtener_logger      

logger = obtener_logger()

# ---------- Inicialización ----------

def crear_tablas():
    """
    Crea la tabla 'stock' si no existe. Informa por consola si ya existía.
    """
    ruta_db = obtener_ruta_db()

    try:
        # 'with' asegura que la conexión se cierre sola, incluso si hay un error
        with sqlite3.connect(str(ruta_db)) as conexion:
            cursor = conexion.cursor()

            # Verificamos si la tabla ya existía ANTES de crearla,
            # para poder informarlo correctamente
            cursor.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='stock'
            """)
            ya_existia = cursor.fetchone() is not None

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS stock (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo TEXT NOT NULL UNIQUE,
                    descripcion TEXT NOT NULL,
                    precio REAL NOT NULL DEFAULT 0,
                    cantidad_caja INTEGER NOT NULL DEFAULT 0,
                    cantidad_stock INTEGER NOT NULL DEFAULT 0,
                    imagen TEXT
                )
            """)
            conexion.commit()

        if ya_existia:
            print("La tabla 'stock' ya existía. No se realizaron cambios.")
        else:
            print("Tabla 'stock' creada correctamente.")

    except sqlite3.Error as e:
        # Error específico de SQLite (por ejemplo, base de datos corrupta)
        logger.error(f"Error al crear la tabla 'stock': {e}")
        raise
    except Exception as e:
        # Cualquier otro error inesperado
        logger.error(f"Error inesperado al crear la tabla: {e}")
        raise


def inicializar_db():
    """
    Función principal para inicializar la base de datos.
    Se llama una vez al iniciar la app (desde main.py).
    """
    try:
        crear_tablas()
        hacer_backup_db()
        print(f"Base de datos lista en: {obtener_ruta_db()}")
    except Exception as e:
        logger.critical(f"No se pudo inicializar la base de datos: {e}")
        raise


# ---------- CRUD de productos ----------

def obtener_productos():
    """
    Devuelve una lista con todos los productos de la tabla stock.
    Cada producto es un diccionario con sus campos.
    """
    ruta_db = obtener_ruta_db()

    try:
        with sqlite3.connect(str(ruta_db)) as conexion:
            conexion.row_factory = sqlite3.Row  # permite acceder a las columnas por nombre
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM stock ORDER BY id")
            filas = cursor.fetchall()

        # Convertimos cada fila en un diccionario normal
        return [dict(fila) for fila in filas]

    except sqlite3.Error as e:
        logger.error(f"Error al obtener productos: {e}")
        raise
    except Exception as e:
        logger.error(f"Error inesperado al obtener productos: {e}")
        raise


def obtener_producto(id=None, codigo=None):
    """
    Devuelve un solo producto, buscándolo por id o por codigo.
    Debe especificarse al menos uno de los dos.
    Devuelve None si no se encuentra.
    """
    if id is None and codigo is None:
        logger.warning("Se llamó a obtener_producto sin especificar id ni codigo.")
        print("Error: hay que especificar 'id' o 'codigo' para buscar un producto.")
        return None

    ruta_db = obtener_ruta_db()

    try:
        with sqlite3.connect(str(ruta_db)) as conexion:
            conexion.row_factory = sqlite3.Row
            cursor = conexion.cursor()

            if id is not None:
                cursor.execute("SELECT * FROM stock WHERE id = ?", (id,))
            else:
                cursor.execute("SELECT * FROM stock WHERE codigo = ?", (codigo,))

            fila = cursor.fetchone()

        return dict(fila) if fila else None

    except sqlite3.Error as e:
        logger.error(f"Error al obtener producto (id={id}, codigo={codigo}): {e}")
        raise
    except Exception as e:
        logger.error(f"Error inesperado al obtener producto (id={id}, codigo={codigo}): {e}")
        raise

def _validar_datos_producto(precio=None, cantidad_caja=None, cantidad_stock=None):
    """
    Valida que los valores numéricos del producto sean correctos.
    Lanza un ValueError si algo no es válido.
    """
    if precio is not None and precio < 0:
        raise ValueError("El precio no puede ser negativo.")
    if cantidad_caja is not None and cantidad_caja < 0:
        raise ValueError("La cantidad por caja no puede ser negativa.")
    if cantidad_stock is not None and cantidad_stock < 0:
        raise ValueError("La cantidad en stock no puede ser negativa.")

def insertar_producto(codigo, descripcion, precio, cantidad_caja=1, cantidad_stock=1, imagen=None):
    """
    Inserta un nuevo producto en la tabla stock.

    - cantidad_caja y cantidad_stock son opcionales, por defecto valen 1.
    - imagen es opcional, por defecto se usa 'default.png'.
    - codigo, descripcion y precio son obligatorios.

    Devuelve el id del producto insertado, o None si falló (ej: código repetido).
    """

    if imagen is None:
        imagen = "default.png"

    try:
        _validar_datos_producto(precio, cantidad_caja, cantidad_stock)
    except ValueError as e:
        logger.warning(f"Datos inválidos al insertar producto '{codigo}': {e}")
        print(f"Error: {e}")
        return None

    ruta_db = obtener_ruta_db()

    try:
        with sqlite3.connect(str(ruta_db)) as conexion:
            cursor = conexion.cursor()
            cursor.execute("""
                INSERT INTO stock (codigo, descripcion, precio, cantidad_caja, cantidad_stock, imagen)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (codigo, descripcion, precio, cantidad_caja, cantidad_stock, imagen))
            conexion.commit()
            nuevo_id = cursor.lastrowid  # id autogenerado del producto recién insertado

        logger.warning(f"Producto creado: codigo='{codigo}', id={nuevo_id}.")
        print(f"Producto '{codigo}' insertado correctamente (id={nuevo_id}).")
        return nuevo_id

    except sqlite3.IntegrityError:
        # Salta cuando se viola el UNIQUE del campo 'codigo'
        logger.warning(f"No se pudo insertar el producto: el código '{codigo}' ya existe.")
        print(f"Error: el código '{codigo}' ya existe.")
        return None
    except sqlite3.Error as e:
        logger.error(f"Error al insertar producto '{codigo}': {e}")
        raise
    except Exception as e:
        logger.error(f"Error inesperado al insertar producto '{codigo}': {e}")
        raise


def editar_producto(id, codigo=None, descripcion=None, precio=None,
                     cantidad_caja=None, cantidad_stock=None, imagen=None):
    """
    Edita un producto existente, identificado por su id.
    Solo se actualizan los campos que se especifiquen (los que queden en None
    no se tocan, así se puede editar un solo campo sin afectar el resto).

    Devuelve True si se actualizó correctamente, False si no se encontró
    el producto o no se especificó ningún campo.
    """
    # Armamos un diccionario con todos los campos posibles
    campos = {
        "codigo": codigo,
        "descripcion": descripcion,
        "precio": precio,
        "cantidad_caja": cantidad_caja,
        "cantidad_stock": cantidad_stock,
        "imagen": imagen,
    }

    campos_a_actualizar = {k: v for k, v in campos.items() if v is not None}

    try:
        _validar_datos_producto(precio, cantidad_caja, cantidad_stock)
    except ValueError as e:
        logger.warning(f"Datos inválidos al editar producto (id={id}): {e}")
        print(f"Error: {e}")
        return False

    if not campos_a_actualizar:
        logger.warning(f"Se llamó a editar_producto (id={id}) sin especificar campos a modificar.")
        print("Error: no se especificó ningún campo para editar.")
        return False

    # Arma dinámicamente algo como: "codigo = ?, precio = ?"
    set_clause = ", ".join(f"{campo} = ?" for campo in campos_a_actualizar)
    valores = list(campos_a_actualizar.values())
    valores.append(id)  # el id va al final, para el WHERE

    ruta_db = obtener_ruta_db()

    try:
        with sqlite3.connect(str(ruta_db)) as conexion:
            cursor = conexion.cursor()
            # Los nombres de columna vienen de nuestro propio diccionario fijo
            # (no de datos del usuario), así que es seguro armar el SQL así.
            # Los valores en cambio siempre van parametrizados con '?'.
            cursor.execute(f"""
                UPDATE stock SET {set_clause} WHERE id = ?
            """, valores)
            conexion.commit()
            filas_afectadas = cursor.rowcount

        if filas_afectadas == 0:
            logger.warning(f"No se encontró producto para editar (id={id}).")
            print("No se encontró ningún producto con ese id.")
            return False
        else:
            logger.warning(f"Producto editado (id={id}): campos modificados: {list(campos_a_actualizar.keys())}.")
            print("Producto editado correctamente.")
            return True

    except sqlite3.IntegrityError:
        # Salta si el nuevo 'codigo' ya lo tiene otro producto
        logger.warning(f"No se pudo editar el producto (id={id}): el código '{codigo}' ya existe en otro producto.")
        print(f"Error: el código '{codigo}' ya existe en otro producto.")
        return False
    except sqlite3.Error as e:
        logger.error(f"Error al editar producto (id={id}): {e}")
        raise
    except Exception as e:
        logger.error(f"Error inesperado al editar producto (id={id}): {e}")
        raise


def eliminar_producto(id=None, codigo=None):
    """
    Elimina un producto de la tabla stock, buscándolo por id o por codigo.
    Debe especificarse al menos uno de los dos parámetros.

    Devuelve True si se eliminó algo, False si no se encontró el producto.
    """
    if id is None and codigo is None:
        logger.warning("Se llamó a eliminar_producto sin especificar id ni codigo.")
        print("Error: hay que especificar 'id' o 'codigo' para eliminar un producto.")
        return False

    ruta_db = obtener_ruta_db()

    try:
        with sqlite3.connect(str(ruta_db)) as conexion:
            cursor = conexion.cursor()

            # Prioriza el id si se especificaron ambos
            if id is not None:
                cursor.execute("DELETE FROM stock WHERE id = ?", (id,))
            else:
                cursor.execute("DELETE FROM stock WHERE codigo = ?", (codigo,))

            conexion.commit()
            filas_afectadas = cursor.rowcount

        if filas_afectadas == 0:
            logger.warning(f"No se encontró producto para eliminar (id={id}, codigo={codigo}).")
            print("No se encontró ningún producto con esos datos.")
            return False
        else:
            logger.warning(f"Producto eliminado (id={id}, codigo={codigo}).")
            print("Producto eliminado correctamente.")
            return True

    except sqlite3.Error as e:
        logger.error(f"Error al eliminar producto (id={id}, codigo={codigo}): {e}")
        raise
    except Exception as e:
        logger.error(f"Error inesperado al eliminar producto (id={id}, codigo={codigo}): {e}")
        raise

# Columnas esperadas en el Excel, en este orden
COLUMNAS_ESPERADAS = ["codigo", "descripcion", "cantidad_por_caja", "cantidad_en_stock", "precio"]

def cargar_productos_excel(ruta_archivo):
    """
    Carga productos masivamente desde un archivo Excel (.xlsx o .xls).

    El archivo debe tener las columnas (en cualquier orden, pero con esos
    nombres exactos en la primera fila):
        codigo | descripcion | cantidad_por_caja | cantidad_en_stock | precio

    - Si hay errores de datos (campo vacío, tipo incorrecto, valor negativo),
      NO se inserta nada y se devuelve la lista de errores encontrados.
    - Si hay productos duplicados dentro del mismo archivo (mismo codigo,
      descripcion y precio), se conserva solo el primero y se informa cuáles
      filas fueron descartadas (esto no bloquea la carga).
    - Si un código ya existe en la base de datos, se informa como error
      (no se inserta ninguno hasta que se corrija el archivo).

    Devuelve un diccionario:
        {
            "exito": True/False,
            "errores": [lista de strings describiendo cada error],
            "duplicados": [lista de strings describiendo duplicados descartados],
            "insertados": cantidad de productos insertados (0 si exito=False)
        }
    """
    resultado = {
        "exito": False,
        "errores": [],
        "duplicados": [],
        "insertados": 0
    }

    # ---------- 1. Abrir el archivo ----------
    try:
        libro = load_workbook(ruta_archivo, data_only=True)
        hoja = libro.active
    except Exception as e:
        logger.error(f"No se pudo abrir el archivo Excel '{ruta_archivo}': {e}")
        resultado["errores"].append(f"No se pudo abrir el archivo: {e}")
        return resultado

    # ---------- 2. Validar encabezados ----------
    primera_fila = [str(celda.value).strip().lower() if celda.value else "" for celda in hoja[1]]

    indices_columnas = {}
    for columna in COLUMNAS_ESPERADAS:
        if columna not in primera_fila:
            resultado["errores"].append(f"Falta la columna obligatoria '{columna}' en la primera fila.")
        else:
            indices_columnas[columna] = primera_fila.index(columna)

    if resultado["errores"]:
        # Si faltan columnas, ni siquiera tiene sentido seguir revisando filas
        logger.warning(f"Carga masiva cancelada por columnas faltantes: {resultado['errores']}")
        return resultado

    # ---------- 3. Leer y validar cada fila ----------
    productos = []  # productos válidos, listos para insertar
    vistos = {}     # para detectar duplicados: clave -> numero de fila

    fila_num = 1  # la fila 1 es el encabezado, los datos arrancan en la 2
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        fila_num += 1

        # Si la fila está completamente vacía, la saltamos sin error
        if all(valor is None for valor in fila):
            continue

        errores_fila = []

        # --- codigo ---
        codigo = fila[indices_columnas["codigo"]]
        if codigo is None or str(codigo).strip() == "":
            errores_fila.append(f"Fila {fila_num}, columna 'codigo': está vacío.")
        else:
            codigo = str(codigo).strip()

        # --- descripcion ---
        descripcion = fila[indices_columnas["descripcion"]]
        if descripcion is None or str(descripcion).strip() == "":
            errores_fila.append(f"Fila {fila_num}, columna 'descripcion': está vacío.")
        else:
            descripcion = str(descripcion).strip()

        # --- cantidad_por_caja ---
        cantidad_caja = fila[indices_columnas["cantidad_por_caja"]]
        if cantidad_caja is None or str(cantidad_caja).strip() == "":
            cantidad_caja = 1  # valor por defecto si viene vacío
        else:
            try:
                cantidad_caja = int(cantidad_caja)
                if cantidad_caja < 0:
                    errores_fila.append(f"Fila {fila_num}, columna 'cantidad_por_caja': no puede ser negativo.")
            except (ValueError, TypeError):
                errores_fila.append(f"Fila {fila_num}, columna 'cantidad_por_caja': debe ser un número entero.")

        # --- cantidad_en_stock ---
        cantidad_stock = fila[indices_columnas["cantidad_en_stock"]]
        if cantidad_stock is None or str(cantidad_stock).strip() == "":
            cantidad_stock = 1  # valor por defecto si viene vacío
        else:
            try:
                cantidad_stock = int(cantidad_stock)
                if cantidad_stock < 0:
                    errores_fila.append(f"Fila {fila_num}, columna 'cantidad_en_stock': no puede ser negativo.")
            except (ValueError, TypeError):
                errores_fila.append(f"Fila {fila_num}, columna 'cantidad_en_stock': debe ser un número entero.")

        # --- precio ---
        precio = fila[indices_columnas["precio"]]
        if precio is None or str(precio).strip() == "":
            errores_fila.append(f"Fila {fila_num}, columna 'precio': está vacío.")
        else:
            try:
                precio = float(precio)
                if precio < 0:
                    errores_fila.append(f"Fila {fila_num}, columna 'precio': no puede ser negativo.")
            except (ValueError, TypeError):
                errores_fila.append(f"Fila {fila_num}, columna 'precio': debe ser un número.")

        if errores_fila:
            resultado["errores"].extend(errores_fila)
            continue  # no seguimos procesando esta fila si ya tiene errores

        # --- Chequeo de duplicados dentro del mismo archivo ---
        clave_duplicado = (codigo, descripcion, precio)
        if clave_duplicado in vistos:
            resultado["duplicados"].append(
                f"Fila {fila_num}: duplicado de la fila {vistos[clave_duplicado]} "
                f"(codigo='{codigo}', descripcion='{descripcion}', precio={precio}). Se descartó."
            )
            continue  # no lo agregamos a productos, nos quedamos con el primero

        vistos[clave_duplicado] = fila_num

        productos.append({
            "codigo": codigo,
            "descripcion": descripcion,
            "cantidad_caja": cantidad_caja,
            "cantidad_stock": cantidad_stock,
            "precio": precio,
            "fila": fila_num
        })

    # ---------- 4. Verificar códigos que ya existen en la base de datos ----------
    for producto in productos:
        existente = obtener_producto(codigo=producto["codigo"])
        if existente is not None:
            resultado["errores"].append(
                f"Fila {producto['fila']}, columna 'codigo': el código '{producto['codigo']}' ya existe en la base de datos."
            )

    # ---------- 5. Si hay errores, no se inserta nada ----------
    if resultado["errores"]:
        logger.warning(
            f"Carga masiva desde '{ruta_archivo}' cancelada. "
            f"{len(resultado['errores'])} error(es) encontrados."
        )
        return resultado

    # ---------- 6. Insertar todos los productos válidos ----------
    try:
        for producto in productos:
            insertar_producto(
                codigo=producto["codigo"],
                descripcion=producto["descripcion"],
                precio=producto["precio"],
                cantidad_caja=producto["cantidad_caja"],
                cantidad_stock=producto["cantidad_stock"]
                # la imagen se deja como 'default.png' (valor por defecto)
            )
        resultado["exito"] = True
        resultado["insertados"] = len(productos)

        logger.warning(
            f"Carga masiva desde '{ruta_archivo}' completada: {len(productos)} producto(s) insertado(s), "
            f"{len(resultado['duplicados'])} duplicado(s) descartado(s)."
        )

    except Exception as e:
        logger.error(f"Error al insertar productos durante la carga masiva: {e}")
        resultado["errores"].append(f"Error al insertar productos: {e}")
        resultado["exito"] = False

    return resultado


# Este bloque solo se ejecuta si corrés este archivo directamente
# (python -m functions.db), no cuando se importa desde otro archivo.
if __name__ == "__main__":
    inicializar_db()