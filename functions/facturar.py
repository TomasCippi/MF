"""
functions/facturar.py

Genera la factura del pedido como Excel, con el formato de la planilla
en papel: PRESUPUESTO/cliente, DIRECCION, R-remito/fecha, tabla de
artículos, totales con descuento y deuda. Se imprimen DOS copias en la
misma hoja (vendedor arriba, cliente abajo), separadas por una línea
gruesa, pensada para entrar en A4 al imprimir.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins

from functions.logger import obtener_logger
from functions.config import obtener_direccion_empresa

logger = obtener_logger()

# ---------- Colores editables ----------
COLOR_NEGRO = "000000"
COLOR_BLANCO = "FFFFFF"
COLOR_GRIS_CLARO = "D9D9D9"

# Bordes: grueso para el contorno de la tabla, mediano para las celdas internas
BORDE_GRUESO = Side(style="thick", color=COLOR_NEGRO)
BORDE_FINO = Side(style="thin", color=COLOR_NEGRO)
BORDE_CELDA = Border(left=BORDE_FINO, right=BORDE_FINO, top=BORDE_FINO, bottom=BORDE_FINO)

# Columnas: A=Artículo(código), B=Descripción, C=Cantidad, D=Precio, E=Total
ANCHO_COLUMNAS = {"A": 14, "B": 42, "C": 12, "D": 14, "E": 16}

FILAS_PRODUCTOS_MAX = 16  # cantidad fija de filas de productos por copia


def generar_factura_excel(ruta_destino, cliente, remito, items, porcentaje_descuento=0, deuda=0):
    """
    Genera el Excel de la factura en 'ruta_destino'. Dibuja dos copias
    (arriba y abajo) separadas por una línea gruesa, en una sola hoja A4.

    items: lista de dicts con codigo, descripcion, precio, cantidad.
    Devuelve True si se generó bien, False si hubo error.
    """
    try:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Factura"

        for col, ancho in ANCHO_COLUMNAS.items():
            hoja.column_dimensions[col].width = ancho

        subtotal = sum(i["precio"] * i["cantidad"] for i in items)
        monto_descuento = subtotal * (porcentaje_descuento / 100)
        total_final = subtotal - monto_descuento + (deuda or 0)

        fecha = datetime.now().strftime("%d-%m-%y")
        direccion = obtener_direccion_empresa()

        ultima_fila_copia_1 = _dibujar_copia(
            hoja, 1, cliente, remito, fecha, direccion, items,
            subtotal, monto_descuento, porcentaje_descuento, deuda, total_final
        )

        # ---------- Línea separadora gruesa entre las dos copias ----------
        fila_separador = ultima_fila_copia_1 + 2
        for col_idx in range(1, 6):
            hoja.cell(row=fila_separador, column=col_idx).border = Border(top=BORDE_GRUESO)
        hoja.row_dimensions[fila_separador].height = 10
        hoja.row_dimensions[ultima_fila_copia_1 + 1].height = 18  # espacio en blanco antes de la línea

        inicio_copia_2 = fila_separador + 2
        hoja.row_dimensions[fila_separador + 1].height = 18  # espacio en blanco después de la línea
        ultima_fila_copia_2 = _dibujar_copia(
            hoja, inicio_copia_2, cliente, remito, fecha, direccion, items,
            subtotal, monto_descuento, porcentaje_descuento, deuda, total_final
        )

        # ---------- Configuración de impresión: ancho a 1 página A4 ----------
        hoja.page_setup.orientation = "portrait"
        hoja.page_setup.paperSize = hoja.PAPERSIZE_A4
        hoja.page_setup.fitToWidth = 1
        hoja.page_setup.fitToHeight = 0
        hoja.sheet_properties.pageSetUpPr.fitToPage = True
        hoja.print_area = f"A1:E{ultima_fila_copia_2}"
        hoja.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3)

        libro.save(ruta_destino)
        logger.info(f"Factura generada en '{ruta_destino}' (remito {remito}, cliente '{cliente}').")
        return True

    except Exception as e:
        logger.error(f"Error al generar la factura: {e}")
        return False


def _dibujar_copia(hoja, fila_inicio, cliente, remito, fecha, direccion, items,
                    subtotal, monto_descuento, porcentaje_descuento, deuda, total_final):
    """
    Dibuja un bloque completo de factura (encabezado + dirección + tabla
    + totales) empezando en 'fila_inicio'. Devuelve la última fila que ocupó.
    """
    fila = fila_inicio

    # ---------- Fila 1: PRESUPUESTO | Cliente (centrado, gris) | R-remito | Fecha ----------
    celda_label = hoja.cell(row=fila, column=1, value="PRESUPUESTO")
    celda_label.font = Font(bold=True, size=11, color=COLOR_NEGRO)
    celda_label.fill = PatternFill("solid", fgColor=COLOR_BLANCO)
    celda_label.alignment = Alignment(horizontal="center", vertical="center")

    hoja.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
    celda_cliente = hoja.cell(row=fila, column=2, value=cliente or "-")
    celda_cliente.font = Font(bold=True, size=13, color=COLOR_NEGRO)
    celda_cliente.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
    celda_cliente.alignment = Alignment(horizontal="center", vertical="center")
    hoja.cell(row=fila, column=3).fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)

    celda_remito = hoja.cell(row=fila, column=4, value=f"R - {remito}")
    celda_remito.font = Font(bold=True, size=12, color=COLOR_NEGRO)
    celda_remito.fill = PatternFill("solid", fgColor=COLOR_BLANCO)
    celda_remito.alignment = Alignment(horizontal="center", vertical="center")

    celda_fecha = hoja.cell(row=fila, column=5, value=fecha)
    celda_fecha.font = Font(bold=True, size=11, color=COLOR_NEGRO)
    celda_fecha.fill = PatternFill("solid", fgColor=COLOR_BLANCO)
    celda_fecha.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 6):
        hoja.cell(row=fila, column=col_idx).border = BORDE_CELDA

    hoja.row_dimensions[fila].height = 24
    fila += 1

    # ---------- Fila 2: DIRECCION | dirección (B:C) | vacío | vacío, todo con fondo gris ----------
    celda_label_dir = hoja.cell(row=fila, column=1, value="DIRECCION")
    celda_label_dir.font = Font(bold=True, size=11, color=COLOR_NEGRO)
    celda_label_dir.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
    celda_label_dir.alignment = Alignment(horizontal="center", vertical="center")

    hoja.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
    celda_direccion = hoja.cell(row=fila, column=2, value=direccion or "-")
    celda_direccion.font = Font(bold=False, size=11, color=COLOR_NEGRO)
    celda_direccion.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
    celda_direccion.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hoja.cell(row=fila, column=3).fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)

    hoja.cell(row=fila, column=4).fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
    hoja.cell(row=fila, column=5).fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)

    for col_idx in range(1, 6):
        hoja.cell(row=fila, column=col_idx).border = BORDE_CELDA

    hoja.row_dimensions[fila].height = 22
    fila += 1

    # ---------- Encabezados de la tabla, fondo negro / letra blanca ----------
    encabezados = ["Artículo", "Descripción", "Cantidad", "Precio", "Total"]
    for col_idx, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila, column=col_idx, value=texto)
        celda.font = Font(bold=True, size=11, color=COLOR_BLANCO)
        celda.fill = PatternFill("solid", fgColor=COLOR_NEGRO)
        celda.border = BORDE_CELDA
        celda.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[fila].height = 20
    fila += 1

    # ---------- Filas de productos: SIEMPRE 16, la columna Artículo
    # (código) queda con fondo gris claro en todas, tengan o no dato ----------
    for i in range(FILAS_PRODUCTOS_MAX):
        if i < len(items):
            item = items[i]
            valores = [
                item["codigo"], item["descripcion"], item["cantidad"],
                item["precio"], item["precio"] * item["cantidad"]
            ]
        else:
            valores = ["", "", "", "", ""]

        for col_idx, valor in enumerate(valores, start=1):
            celda = hoja.cell(row=fila, column=col_idx, value=valor if valor != "" else None)
            celda.border = BORDE_CELDA

            if col_idx == 1:
                celda.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
                celda.font = Font(bold=True, size=10)
                celda.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                celda.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 3:
                celda.font = Font(bold=True, size=11)
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if isinstance(valor, (int, float)):
                    celda.number_format = '"$"#,##0.00'
                celda.alignment = Alignment(horizontal="center", vertical="center")
        fila += 1

    # ---------- Totales: etiqueta ocupa A:D (fondo gris), valor en E ----------
    def _fila_total(texto, valor, tamano=11):
        nonlocal fila
        hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)

        celda_txt = hoja.cell(row=fila, column=1, value=texto)
        celda_txt.font = Font(bold=True, size=tamano)
        celda_txt.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        celda_txt.alignment = Alignment(horizontal="right", vertical="center")

        for col_idx in range(1, 5):
            hoja.cell(row=fila, column=col_idx).border = BORDE_CELDA
            hoja.cell(row=fila, column=col_idx).fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)

        celda_valor = hoja.cell(row=fila, column=5, value=valor)
        celda_valor.font = Font(bold=True, size=tamano)
        celda_valor.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        celda_valor.number_format = '"$"#,##0.00' if isinstance(valor, (int, float)) else "General"
        celda_valor.alignment = Alignment(horizontal="center", vertical="center")
        celda_valor.border = BORDE_CELDA

        hoja.row_dimensions[fila].height = 20
        fila += 1

    _fila_total("TOTAL", subtotal)

    if monto_descuento > 0:
        # ---------- Fila de descuento: A y B vacíos (gris), C = "Descuento",
        # D = porcentaje (10%, 15%, etc.), E = monto descontado en $ ----------
        for col_idx in range(1, 3):
            celda_vacia = hoja.cell(row=fila, column=col_idx)
            celda_vacia.border = BORDE_CELDA
            celda_vacia.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)

        celda_label_desc = hoja.cell(row=fila, column=3, value="Descuento")
        celda_label_desc.font = Font(bold=True, size=11, color=COLOR_NEGRO)
        celda_label_desc.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        celda_label_desc.border = BORDE_CELDA
        celda_label_desc.alignment = Alignment(horizontal="right", vertical="center")

        celda_porcentaje = hoja.cell(row=fila, column=4, value=f"{porcentaje_descuento:g}%")
        celda_porcentaje.font = Font(bold=True, size=11, color=COLOR_NEGRO)
        celda_porcentaje.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        celda_porcentaje.border = BORDE_CELDA
        celda_porcentaje.alignment = Alignment(horizontal="center", vertical="center")

        celda_monto_desc = hoja.cell(row=fila, column=5, value=monto_descuento)
        celda_monto_desc.font = Font(bold=True, size=11, color=COLOR_NEGRO)
        celda_monto_desc.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        celda_monto_desc.number_format = '"$"#,##0.00'
        celda_monto_desc.border = BORDE_CELDA
        celda_monto_desc.alignment = Alignment(horizontal="center", vertical="center")

        hoja.row_dimensions[fila].height = 20
        fila += 1

    if deuda and deuda > 0:
        _fila_total("SALDO ANTERIOR", deuda)

    _fila_total("SALDO ACTUAL", total_final, tamano=13)

    return fila - 1
